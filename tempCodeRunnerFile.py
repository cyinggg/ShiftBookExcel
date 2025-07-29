# main.py
import telebot
import os
import tempfile # Export summary log to excel
from keep_alive import keep_alive  # For Replit uptime

from datetime import datetime, timedelta
import threading
import time
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InputFile
import pytz

from openpyxl import load_workbook, Workbook

# Load your token from environment
from dotenv import load_dotenv

# === Load Telegram token ===
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
bot = telebot.TeleBot(TOKEN)

# === Telegram Group IDs ===
# G1 receive all, G2 cancel and rebook
GROUP_1_CHAT_ID = -1002635519712
#GROUP_2_CHAT_ID = 

# === Track cancelled shifts to catch rebook events ===
cancelled_shifts = set()

def load_cancelled_shifts():
    """Populate cancelled_shifts from cancellations.xlsx on startup."""
    try:
        wb = load_workbook(CANCELLATIONS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, _, date, shift = row[:4]
            cancelled_shifts.add((date, shift))
    except FileNotFoundError:
        pass

# === Excel file paths ===
BOOKINGS_FILE = "bookings.xlsx"
STUDENTS_FILE = "students.xlsx"
CANCELLATIONS_FILE = "cancellations.xlsx"
SUMMARY_FILE = "summary.xlsx"

# Run loader
load_cancelled_shifts()

# === Shift definitions ===
SHIFT_OPTIONS = {
    "Morning": ("09:00", "12:00"),
    "Afternoon": ("14:00", "18:00"),
    "Night": ("18:00", "22:00")
}

# Cache login state using student ID as key
logged_in_users = {}

# === Helper function to check if a student is valid (based on students.xlsx) ===
def is_valid_student(student_id, name):
    wb = load_workbook("students.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(student_id) and row[1].lower() == name.lower():
            return True
    return False

# === Session cache ===
logged_in_users = {}

def notify_group1(msg): bot.send_message(GROUP_1_CHAT_ID, msg)
def notify_group2(msg): bot.send_message(GROUP_2_CHAT_ID, msg)

# Help function to retrieve studentID properly
# Helper to get student_id from session
def get_student_info(student_id):
    if not os.path.exists(STUDENTS_FILE):
        return None
    wb = load_workbook(STUDENTS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(student_id):
            return row # Make sure it returns full row
    return None

# Help function for writing to summary log
def log_to_summary(action, sid, name, date, shift):
    timestamp = datetime.now(pytz.timezone("Asia/Singapore")).strftime("%Y-%m-%d %H:%M:%S")
    if not os.path.exists(SUMMARY_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Timestamp","Action","StudentID","Name","Date","Shift","LIC","LIC Verified"])
    else:
        wb = load_workbook(SUMMARY_FILE)
        ws = wb.active
    ws.append([timestamp, action, sid, name, date, shift, "N/A", "N/A"])
    wb.save(SUMMARY_FILE)

# Bot manual
# Manual/help message to guide user
def send_manual(chat_id):
    manual = ("*User Manual*\n\n"
              "*Login*: /login\n"
              "*Reserve*: /book new shift\n"
              "*Cancel*: /cancel booked shift\n"
              "*MyShifts*: /mybookings view upcoming booked shift\n"
              "*Summary*: PODs can use /summary_log to export all bookings.\n\n"
              "*Shift Rules:*\n"
              "• Max 4/2 shifts/week (unless within 48 hours / 5 days).\n"
              "• Night shifts only for selected SCs (Wed/Thu).\n"
              "• You can book Morning + Afternoon, but *not* Afternoon + Night.\n\n"
              "If you encounter issues, please drop a text in SC chat."
              )
    bot.send_message(chat_id, manual, parse_mode='Markdown')

def is_logged_in(uid): return uid in logged_in_users
def get_student_id_from_session(uid): return logged_in_users.get(uid, {}).get("student_id")

def get_user_bookings(student_id):
    if not os.path.exists(BOOKINGS_FILE): return []
    wb = load_workbook(BOOKINGS_FILE)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid, _, date_str, shift = row
        if str(sid) == str(student_id):
            out.append({"date": datetime.strptime(date_str, "%Y-%m-%d").date(), "shift": shift})
    return out

# === Handlers: manual, start, reserve, cancel, mybookings, summary_log ===
# /Manual commond handler
@bot.message_handler(commands=['manual'])
def manual_handler(msg):
    send_manual(msg.chat.id)

# /Start commond handler
@bot.message_handler(commands=['start'])
def start_handler(msg):
    bot.send_message(msg.chat.id, "Enter your Student ID:")
    bot.register_next_step_handler(msg, get_student_id)

def get_student_id(msg):
    sid = msg.text.strip()
    if not (sid.isdigit() and len(sid)==7):
        bot.send_message(msg.chat.id, "Invalid ID. Use /start again.")
        return
    bot.send_message(msg.chat.id, "Enter your name:")
    bot.register_next_step_handler(msg, get_student_name, sid)

def get_student_name(msg, sid):
    name = msg.text.strip()
    if not name.isalpha():
        bot.send_message(msg.chat.id, "Invalid name. Use /start again.")
        return
    info = get_student_info(sid)
    if info and info[1].lower()==name.lower():
        logged_in_users[msg.from_user.id] = {"student_id": sid, "name": name, "is_admin": len(info)>4 and info[4]==1}
        bot.send_message(msg.chat.id, f"Login success, {name}!")
        send_manual(msg.chat.id)
    else:
        bot.send_message(msg.chat.id, "Invalid credentials. Use /start again.")

#---Booking---
# /reserve command handler
@bot.message_handler(commands=['reserve'])
def reserve_handler(message):
    if not is_logged_in(message.from_user.id):
        bot.send_message(message.chat.id, "You are not logged in. Use /start.")
        return
    user_data = logged_in_users.get(message.from_user.id)
    if not user_data or "student_id" not in user_data:
        bot.send_message(message.chat.id, "Session expired or invalid. Use /start.")
        return

    student_id = user_data["student_id"]
    print(f"[DEBUG] reserve_handler: student_id={student_id}")

    bot.send_message(message.chat.id, "Enter date to book (YYYY-MM-DD):")
    bot.register_next_step_handler(message, handle_date_selection, student_id)

# handle date input and show available shifts
def handle_date_selection(message, student_id):
    print(f"[DEBUG] handle_date_selection: student_id={student_id}, message={message.text}")

    try:
        selected_date = datetime.strptime(message.text.strip(), "%Y-%m-%d").date()
        today = datetime.now(pytz.timezone("Asia/Singapore")).date()
        # Check if it's for next month, and only open 5 days in advance
        if selected_date.month > today.month or selected_date.year > today.year:
            delta_days = (selected_date - today).days
            if delta_days > 5 or datetime.now(pytz.timezone("Asia/Singapore")).hour < 18:
                bot.send_message(message.chat.id, "Next month's booking opens only 5 days before at 6PM SG time.")
                return

        # Night shift check (only Wed/Thu)
        day_of_week = selected_date.weekday()
        is_night_allowed = get_student_info(student_id)[3] == 1

        available_shifts = []
        bookings = get_user_bookings(student_id)
        booked_today = [b['shift'] for b in bookings if b['date'] == selected_date]
        shift_counts = {
            "Morning": 1,
            "Afternoon": 2,
            "Night": 2 if day_of_week in [2, 3] and is_night_allowed else 0
        }

        # Load existing bookings
        wb = load_workbook(BOOKINGS_FILE)
        ws = wb.active
        current_count = {k: 0 for k in SHIFT_OPTIONS}
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, _, date_str, shift = row
            if date_str == selected_date.strftime("%Y-%m-%d"):
                current_count[shift] += 1

        for shift in SHIFT_OPTIONS:
            if current_count[shift] < shift_counts[shift]:
                if shift == "Night" and not is_night_allowed:
                    continue
                if shift == "Night" and "Afternoon" in booked_today:
                    continue
                available_shifts.append(shift)

        if not available_shifts:
            bot.send_message(message.chat.id, "No available shifts for {selected_date}.")
            return

        markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        for shift in available_shifts:
            markup.add(KeyboardButton(shift))

        bot.send_message(message.chat.id, "Select a shift:", reply_markup=markup)
        bot.register_next_step_handler(message, finalize_booking, student_id, selected_date)

    except ValueError:
        bot.send_message(message.chat.id, "Invalid date format. Please try `/reserve` again.")

# Finalize booking and save to excel
def finalize_booking(message, student_id, selected_date):
    chosen_shift = message.text.strip()
    if chosen_shift not in SHIFT_OPTIONS:
        bot.send_message(message.chat.id, "Invalid shift.")
        return

    bookings = get_user_bookings(student_id)
    student_info = get_student_info(student_id)
    name = student_info[1]

    special_user = len(student_info) > 5 and student_info[5] == 1

    week_start = selected_date - timedelta(days=selected_date.weekday())
    week_end = week_start + timedelta(days=6)
    now = datetime.now(pytz.timezone("Asia/Singapore"))

    weekly_bookings = [b for b in bookings if week_start <= b["date"] <= week_end]
    days_ahead = (selected_date - now.date()).days
    within_5_days = days_ahead < 5
    within_48_hours = (datetime.combine(selected_date, datetime.min.time()) - now).total_seconds() < 48 * 3600

    if special_user:
        if len(weekly_bookings) >= 2 and not within_48_hours:
            bot.send_message(message.chat.id, "Max 2 shifts/week for your account (unless within 48h).")
            return
    else:
        if len(weekly_bookings) >= 4 and not within_5_days:
            bot.send_message(message.chat.id, "Max 4 shifts/week (unless within next 5 days).")
            return

    # Save booking
    wb = load_workbook(BOOKINGS_FILE)
    ws = wb.active
    ws.append([student_info[0], name, selected_date.strftime("%Y-%m-%d"), chosen_shift])
    wb.save(BOOKINGS_FILE)

    log_to_summary("BOOKED", student_info[0], name, selected_date.strftime("%Y-%m-%d"), chosen_shift)
    bot.send_message(message.chat.id, f"Booking confirmed for {selected_date} ({chosen_shift})!")
    send_manual(message.chat.id)

    # Notify groups
    notify_group1(f"*Booked:* {name} ({student_id}) on {selected_date} [{chosen_shift}]", parse_mode='Markdown')

    # If previously cancelled
    if (selected_date.strftime("%Y-%m-%d"), chosen_shift) in cancelled_shifts:
        notify_group2(f"*Rebooked Cancelled Shift:* {name} ({student_id}) on {selected_date} [{chosen_shift}]", parse_mode='Markdown')
        cancelled_shifts.discard((selected_date.strftime("%Y-%m-%d"), chosen_shift))


# Cancel booked shift
@bot.message_handler(commands=['cancel'])
def cancel_handler(message):
    if not is_logged_in(message.from_user.id):
        bot.send_message(message.chat.id, "You are not logged in.")
        return

    student_id = get_student_id_from_session(message.from_user.id)
    bookings = get_user_bookings(student_id)
    future = [b for b in bookings if b["date"] >= datetime.today().date()]

    if not future:
        bot.send_message(message.chat.id, "No future bookings to cancel.")
        return

    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    options = {}
    for b in future:
        label = f"{b['date']} - {b['shift']}"
        markup.add(KeyboardButton(label))
        options[label] = b

    bot.send_message(message.chat.id, "Select booking to cancel:", reply_markup=markup)
    bot.register_next_step_handler_by_chat_id(message.chat.id, confirm_cancel, student_id, options)

def confirm_cancel(message, student_id, booking_map):
    selected = message.text.strip()
    if selected not in booking_map:
        bot.send_message(message.chat.id, "Invalid selection.")
        return

    b = booking_map[selected]
    date_str = b['date'].strftime("%Y-%m-%d")
    shift = b['shift']
    student = get_student_info(student_id)
    name = student[1]

    # Delete from bookings.xlsx
    wb = load_workbook(BOOKINGS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if (str(row[0].value) == student_id and row[2].value == date_str and row[3].value == shift):
            ws.delete_rows(row[0].row)
            break
    wb.save(BOOKINGS_FILE)

    # Append to cancellations.xlsx
    log_wb = load_workbook(CANCELLATIONS_FILE) if os.path.exists(CANCELLATIONS_FILE) else Workbook()
    log_ws = log_wb.active
    if log_ws.max_row == 1:
        log_ws.append(["Timestamp", "StudentID", "Name", "Date", "Shift", "LIC", "LIC Verified"])
    log_ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), student_id, name, date_str, shift, "N/A", "N/A"])
    log_wb.save(CANCELLATIONS_FILE)

    # Add to cancelled_shifts set
    cancelled_shifts.add((date_str, shift))
    log_to_summary("CANCELLED", student_id, name, date_str, shift)

    bot.send_message(message.chat.id, f"Booking on {date_str} ({shift}) cancelled.")
    send_manual(message.chat.id)

    # Notify groups
    notify_group1(f"*Cancelled:* {name} ({student_id}) on {date_str} [{shift}]", parse_mode='Markdown')
    notify_group2(f"*Shift Cancelled:* {name} ({student_id}) on {date_str} [{shift}]", parse_mode='Markdown')


#  User booked summary
@bot.message_handler(commands=['mybookings'])
def my_bookings_handler(message):
    if not is_logged_in(message.from_user.id):
        bot.send_message(message.chat.id, "You are not logged in. Use /start to log in.")
        return

    student_id = get_student_id_from_session(message.from_user.id)
    all_bookings = get_user_bookings(student_id)

    if not all_bookings:
        bot.send_message(message.chat.id, "You have no bookings.")
        return

    today = datetime.now(pytz.timezone("Asia/Singapore")).date()

    # Filter for today and future dates
    future_bookings = [b for b in all_bookings if b['date'] >= today]

    if not future_bookings:
        bot.send_message(message.chat.id, "You have no upcoming bookings.")
        return

    # Sort by date
    sorted_bookings = sorted(future_bookings, key=lambda x: x['date'])

    message_lines = ["*Your Upcoming Shifts:*"]
    for b in sorted_bookings:
        shift_time = ""
        if b['shift'].lower() == "morning":
            shift_time = "(9AM–12PM)"
        elif b['shift'].lower() == "afternoon":
            shift_time = "(2PM–6PM)"
        elif b['shift'].lower() == "night":
            shift_time = "(6PM–10PM)"
        else:
            shift_time = ""  # fallback in case of invalid entry

        message_lines.append(f"• {b['date']} - {b['shift'].capitalize()} {shift_time}")

    response = "\n".join(message_lines)
    bot.send_message(message.chat.id, response, parse_mode="Markdown")


# Only allow admin to access to summary log
@bot.message_handler(commands=['summary_log'])
def summary_log_handler(message):
    user = logged_in_users.get(message.from_user.id)
    if not user or not user.get("is_admin"):
        bot.send_message(message.chat.id, "Unauthorized.")
        return

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        summary_path = tmp.name

    wb = Workbook()

    # ------------------ Bookings Sheet ------------------
    if os.path.exists(BOOKINGS_FILE):
        booking_ws = load_workbook(BOOKINGS_FILE).active
        main_sheet = wb.active
        main_sheet.title = "Bookings"
        for row in booking_ws.iter_rows(values_only=True):
            main_sheet.append(row)
    else:
        wb.active.append(["No bookings found."])

    # ------------------ Cancellations Sheet ------------------
    if os.path.exists(CANCELLATIONS_FILE):
        cancel_ws = load_workbook(CANCELLATIONS_FILE).active
        sheet = wb.create_sheet("Cancellations")
        for row in cancel_ws.iter_rows(values_only=True):
            sheet.append(row)

    # Save final summary Excel file
    wb.save(summary_path)

    # Send as Telegram document
    with open(summary_path, "rb") as f:
        bot.send_document(message.chat.id, InputFile(f, filename="ShiftSummary.xlsx"))

    # Cleanup temp file
    os.remove(summary_path)

# Auto notification of the upcoming shift one hour in advance
def shift_reminder_loop():
    while True:
        now = datetime.now(pytz.timezone("Asia/Singapore"))
        today = now.date()

        # Load today's bookings
        if not os.path.exists(BOOKINGS_FILE):
            time.sleep(60)
            continue

        wb = load_workbook(BOOKINGS_FILE)
        ws = wb.active

        for shift_name, (start_str, _) in SHIFT_OPTIONS.items():
            shift_start_time = datetime.strptime(start_str, "%H:%M").time()
            shift_datetime = datetime.combine(today, shift_start_time)
            shift_datetime = pytz.timezone("Asia/Singapore").localize(shift_datetime)

            # Notify exactly 1 hour before the shift
            time_diff = (shift_datetime - now).total_seconds()
            if 3540 <= time_diff <= 3660:  # ~1 hour ±1 minute window
                students_in_shift = [
                    (row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[2] == today.strftime("%Y-%m-%d") and row[3] == shift_name
                ]

                if students_in_shift:
                    msg_lines = [f"*Shift Reminder: {shift_name} ({start_str})*", f"*Date:* {today.strftime('%Y-%m-%d')}"]
                    for sid, name in students_in_shift:
                        msg_lines.append(f"{name} (ID: {sid})")
                    message = "\n".join(msg_lines)

                    # Notify all logged-in users in the shift
                    for user_id, info in logged_in_users.items():
                        if (info["student_id"], info["name"]) in students_in_shift:
                            bot.send_message(user_id, message, parse_mode='Markdown')

        # Sleep 60 seconds before next check
        time.sleep(60)

# Run 24/7
from keep_alive import keep_alive
if __name__ == "__main__":
    keep_alive()
    print("Bot is running.")
    bot.polling(non_stop=True)
