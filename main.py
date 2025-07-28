# main.py
import telebot
import os
import tempfile

from datetime import datetime, timedelta
import threading
import time
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InputFile
import pytz

# Export summary log to excel
import tempfile
from openpyxl import load_workbook, Workbook

# Load your token from environment
from dotenv import load_dotenv
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
bot = telebot.TeleBot(TOKEN)

# Excel file paths
STUDENT_FILE = "students.xlsx"
BOOKINGS_FILE = "bookings.xlsx"
CANCELLATIONS_FILE = "cancellations.xlsx"

# Define shifts
SHIFT_OPTIONS = {
    "Morning": ("09:00", "12:00"),
    "Afternoon": ("14:00", "18:00"),
    "Night": ("18:00", "22:00")
}

# Cache login state using student ID as key
# {student_id: {"name": name, "chat_id": chat_id}}
logged_in_users = {}

# Group chat notification
GROUP_CHAT_ID = -1002635519712
def notify_group(message_text):
    try:
        bot.send_message(GROUP_CHAT_ID, message_text)
    except Exception as e:
        print(f"[ERROR] Failed to notify group: {e}")


# Help function to retrieve studentID properly
# Helper to get student_id from session
def get_student_info(student_id):
    if not os.path.exists(STUDENT_FILE):
        return None
    wb = load_workbook(STUDENT_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(student_id):
            return row  # Make sure it returns full row
    return None


# Help function for writing to summary log
def log_to_summary(action_type, student_id, name, date, shift, lic="N/A", lic_verified="N/A"):
    summary_file = "summary.xlsx"
    timestamp = datetime.now(pytz.timezone("Asia/Singapore")).strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(summary_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Timestamp", "Action", "StudentID", "Name", "Date", "Shift", "LIC", "LIC Verified"])
    else:
        wb = load_workbook(summary_file)
        ws = wb.active

    ws.append([timestamp, action_type, student_id, name, date, shift, lic, lic_verified])
    wb.save(summary_file)

# Bot manual
# Manual/help message to guide user
def send_manual(chat_id):
    manual = (
        "*User Manual*\n\n"
        "*Login*: Use /start and provide your Student ID and Name.\n"
        "*Reserve*: Use /reserve to book a shift.\n"
        "*Cancel*: Use /cancel to cancel an upcoming booking.\n"
        "*MyShifts*: Use /mybookings to view your upcoming bookings.\n"
        "*Summary*: PODs can use /summary_log to export all bookings.\n\n"
        "*Shift Rules:*\n"
        "• Max 4/2 shifts/week (unless within 48 hours / 5 days).\n"
        "• Night shifts only for selected SCs (Wed/Thu).\n"
        "• You can book Morning + Afternoon, but *not* Afternoon + Night.\n\n"
        "If you encounter issues, please drop a text in SC chat."
    )
    bot.send_message(chat_id, manual, parse_mode='Markdown')

# Check login
def is_logged_in(telegram_user_id):
    return telegram_user_id in logged_in_users

def get_student_id_from_session(telegram_user_id):
    user_data = logged_in_users.get(telegram_user_id)
    return user_data.get("student_id") if user_data else None

def get_user_bookings(student_id):
    if not os.path.exists(BOOKINGS_FILE):
        return []
    wb = load_workbook(BOOKINGS_FILE)
    ws = wb.active
    bookings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid, _, date_str, shift = row
        if str(sid) == str(student_id):
            bookings.append({"date": datetime.strptime(date_str, "%Y-%m-%d").date(), "shift": shift})
    return bookings


# /Manual commond handler
@bot.message_handler(commands=['manual'])
def manual_handler(message):
    send_manual(message.chat.id)

# /Start commond handler
@bot.message_handler(commands=['start'])
def handle_start(message):
    bot.send_message(message.chat.id, "Welcome to ShiftBookTeleBot!\n Please enter your Student ID:")
    bot.register_next_step_handler(message, get_student_id)

def get_student_id(message):
    student_id = message.text.strip()
    
    # Validate: Must be 7 digits
    if not student_id.isdigit() or len(student_id) != 7:
        bot.send_message(message.chat.id, "Invalid. Enter your 7-digit Student ID. Please try /start again.")
        return

    bot.send_message(message.chat.id, "Enter your name:")
    bot.register_next_step_handler(message, get_student_name, student_id)

def get_student_name(message, student_id):
    name = message.text.strip()

    # Validate: Alphabet characters only
    if not name.isalpha():
        bot.send_message(message.chat.id, "Invalid. Enter you name (alphabet characters only). Try /start again.")
        return

    student_info = get_student_info(student_id)
    if student_info and student_info[1].lower() == name.lower():
        is_admin = len(student_info) > 4 and student_info[4] == 1
        logged_in_users[message.from_user.id] = {
            "student_id": student_id,
            "name": name,
            "is_admin": is_admin
        }
        bot.send_message(message.chat.id, f"Login successful, {name}!")
        send_manual(message.chat.id)
    else:
        bot.send_message(message.chat.id, "Invalid credentials. Please try /start again.")

#---Booking---

# /reserve command handler
@bot.message_handler(commands=['reserve'])
def reserve_handler(message):
    if not is_logged_in(message.from_user.id):
        bot.send_message(message.chat.id, "You are not logged in. Use /start.")
        return
    user_data = logged_in_users.get(message.from_user.id)
    if not user_data or "student_id" not in user_data:
        bot.send_message(message.chat.id, "Session expired or invalid. Use /start.")
        return

    student_id = user_data["student_id"]
    print(f"[DEBUG] reserve_handler: student_id={student_id}")

    bot.send_message(message.chat.id, "Enter date to book (YYYY-MM-DD):")
    bot.register_next_step_handler(message, handle_date_selection, student_id)

# handle date input and show available shifts
def handle_date_selection(message, student_id):
    print(f"[DEBUG] handle_date_selection: student_id={student_id}, message={message.text}")

    try:
        selected_date = datetime.strptime(message.text.strip(), "%Y-%m-%d").date()
        today = datetime.now(pytz.timezone("Asia/Singapore")).date()
        # Check if it's for next month, and only open 5 days in advance
        if selected_date.month > today.month or selected_date.year > today.year:
            delta_days = (selected_date - today).days
            if delta_days > 5 or datetime.now(pytz.timezone("Asia/Singapore")).hour < 18:
                bot.send_message(message.chat.id, "Next month's booking opens only 5 days before at 6PM SG time.")
                return

        # Night shift check (only Wed/Thu)
        day_of_week = selected_date.weekday()
        is_night_allowed = get_student_info(student_id)[3] == 1

        available_shifts = []
        bookings = get_user_bookings(student_id)
        booked_today = [b['shift'] for b in bookings if b['date'] == selected_date]
        shift_counts = {
            "Morning": 1,
            "Afternoon": 2,
            "Night": 2 if day_of_week in [2, 3] and is_night_allowed else 0
        }

        # Load existing bookings
        wb = load_workbook(BOOKINGS_FILE)
        ws = wb.active
        current_count = {k: 0 for k in SHIFT_OPTIONS}
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, _, date_str, shift = row
            if date_str == selected_date.strftime("%Y-%m-%d"):
                current_count[shift] += 1

        for shift in SHIFT_OPTIONS:
            if current_count[shift] < shift_counts[shift]:
                if shift == "Night" and not is_night_allowed:
                    continue
                if shift == "Night" and "Afternoon" in booked_today:
                    continue
                available_shifts.append(shift)

        if not available_shifts:
            bot.send_message(message.chat.id, "No available shifts for {selected_date}.")
            return

        markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        for shift in available_shifts:
            markup.add(KeyboardButton(shift))

        bot.send_message(message.chat.id, "Select a shift:", reply_markup=markup)
        bot.register_next_step_handler(message, finalize_booking, student_id, selected_date)

    except ValueError:
        bot.send_message(message.chat.id, "Invalid date format. Please try `/reserve` again.")

# Finalize booking and save to excel
def finalize_booking(message, student_id, selected_date):
    chosen_shift = message.text.strip()
    if chosen_shift not in SHIFT_OPTIONS:
        bot.send_message(message.chat.id, "Invalid shift.")
        return
    
    bookings = get_user_bookings(student_id)
    student_info = get_student_info(student_id)

    special_user = False
    if len(student_info) > 5 and student_info[5] == 1:
        special_user = True

    week_start = selected_date - timedelta(days=selected_date.weekday())
    week_end = week_start + timedelta(days=6)
    now = datetime.now(pytz.timezone("Asia/Singapore"))

    # Count weekly bookings (excluding bookings within leniency period)
    weekly_bookings = [
        b for b in bookings
        if week_start <= b["date"] <= week_end
    ]

    # Determine leniency window
    days_ahead = (selected_date - now.date()).days
    within_5_days = days_ahead < 5
    within_48_hours = (datetime.combine(selected_date, datetime.min.time()) - now).total_seconds() < 48 * 3600

    if special_user:
        if len(weekly_bookings) >= 2 and not within_48_hours:
            bot.send_message(message.chat.id, "Max 2 shifts/week for your account (unless within 48h).")
            return
    else:
        if len(weekly_bookings) >= 4 and not within_5_days:
            bot.send_message(message.chat.id, "Max 4 shifts/week (unless within next 5 days).")
            return


    wb = load_workbook(BOOKINGS_FILE)
    ws = wb.active
    student_info = get_student_info(student_id)
    ws.append([student_info[0], student_info[1], selected_date.strftime("%Y-%m-%d"), chosen_shift])
    wb.save(BOOKINGS_FILE)
    # Log to summary
    log_to_summary("BOOKED", student_info[0], student_info[1], selected_date.strftime("%Y-%m-%d"), chosen_shift)

    bot.send_message(message.chat.id, f"Booking confirmed for {selected_date} ({chosen_shift})!")
    send_manual(message.chat.id)

    # Check if this shift was previously cancelled
    rebooked_shift = False
    if os.path.exists(CANCELLATIONS_FILE):
        cancel_wb = load_workbook(CANCELLATIONS_FILE)
        cancel_ws = cancel_wb.active
        for row in cancel_ws.iter_rows(min_row=2, values_only=True):
            if str(row[2]) == selected_date.strftime("%Y-%m-%d") and row[3] == chosen_shift:
                rebooked_shift = True
                break

    # Only notify group if this shift was previously cancelled
    if rebooked_shift:
        notify_group(f"*Rebooked Shift!*\n {student_info[1]} ({student_info[0]})\n {selected_date.strftime('%Y-%m-%d')}\n {chosen_shift}")


# Cancel booked shift
@bot.message_handler(commands=['cancel'])
def cancel_handler(message):
    if not is_logged_in(message.from_user.id):
        bot.send_message(message.chat.id, "You are not logged in.")
        return

    student_id = get_student_id_from_session(message.from_user.id)
    bookings = get_user_bookings(student_id)
    future = [b for b in bookings if b["date"] >= datetime.today().date()]

    if not future:
        bot.send_message(message.chat.id, "No future bookings to cancel.")
        return

    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    options = {}
    for b in future:
        label = f"{b['date']} - {b['shift']}"
        markup.add(KeyboardButton(label))
        options[label] = b

    bot.send_message(message.chat.id, "Select booking to cancel:", reply_markup=markup)
    bot.register_next_step_handler_by_chat_id(message.chat.id, confirm_cancel, student_id, options)

def confirm_cancel(message, student_id, booking_map):
    selected = message.text.strip()
    if selected not in booking_map:
        bot.send_message(message.chat.id, "Invalid selection.")
        return

    b = booking_map[selected]
    wb = load_workbook(BOOKINGS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if (str(row[0].value) == student_id and row[2].value == b['date'].strftime("%Y-%m-%d") and row[3].value == b['shift']):
            ws.delete_rows(row[0].row)
            wb.save(BOOKINGS_FILE)
            break

    log_wb = load_workbook(CANCELLATIONS_FILE) if os.path.exists(CANCELLATIONS_FILE) else Workbook()
    log_ws = log_wb.active
    timestamp = datetime.now(pytz.timezone("Asia/Singapore")).strftime("%Y-%m-%d %H:%M:%S")
    log_ws.append([timestamp, student[0], student[1], b['date'].strftime("%Y-%m-%d"), b['shift'], "N/A", "N/A"])
    student = get_student_info(student_id)
    log_ws.append([student[0], student[1], b['date'].strftime("%Y-%m-%d"), b['shift'], datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    log_wb.save(CANCELLATIONS_FILE)
    # Log to summary
    log_to_summary("CANCELLED", student[0], student[1], b['date'].strftime("%Y-%m-%d"), b['shift'])


    bot.send_message(message.chat.id, f"Booking on {b['date']} ({b['shift']}) cancelled.")
    send_manual(message.chat.id)

    # Group notification
    notify_group(f"*Shift Cancelled!*\n {student[1]} ({student[0]})\n {b['date']} - {b['shift']}")

#  User booked summary
@bot.message_handler(commands=['mybookings'])
def my_bookings_handler(message):
    if not is_logged_in(message.from_user.id):
        bot.send_message(message.chat.id, "You are not logged in. Use /start to log in.")
        return

    student_id = get_student_id_from_session(message.from_user.id)
    all_bookings = get_user_bookings(student_id)

    if not all_bookings:
        bot.send_message(message.chat.id, "You have no bookings.")
        return

    today = datetime.now(pytz.timezone("Asia/Singapore")).date()

    # Filter for today and future dates
    future_bookings = [b for b in all_bookings if b['date'] >= today]

    if not future_bookings:
        bot.send_message(message.chat.id, "You have no upcoming bookings.")
        return

    # Sort by date
    sorted_bookings = sorted(future_bookings, key=lambda x: x['date'])

    message_lines = ["*Your Upcoming Shifts:*"]
    for b in sorted_bookings:
        shift_time = ""
        if b['shift'].lower() == "morning":
            shift_time = "(9AM–12PM)"
        elif b['shift'].lower() == "afternoon":
            shift_time = "(2PM–6PM)"
        elif b['shift'].lower() == "night":
            shift_time = "(6PM–10PM)"
        else:
            shift_time = ""  # fallback in case of invalid entry

        message_lines.append(f"• {b['date']} - {b['shift'].capitalize()} {shift_time}")

    response = "\n".join(message_lines)
    bot.send_message(message.chat.id, response, parse_mode="Markdown")


# Only allow admin to access to summary log
@bot.message_handler(commands=['summary_log'])
def summary_log_handler(message):
    user = logged_in_users.get(message.from_user.id)
    if not user or not user.get("is_admin"):
        bot.send_message(message.chat.id, "Unauthorized.")
        return

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        summary_path = tmp.name

    wb = Workbook()

    # ------------------ Bookings Sheet ------------------
    if os.path.exists(BOOKINGS_FILE):
        booking_ws = load_workbook(BOOKINGS_FILE).active
        main_sheet = wb.active
        main_sheet.title = "Bookings"
        for row in booking_ws.iter_rows(values_only=True):
            main_sheet.append(row)
    else:
        wb.active.append(["No bookings found."])

    # ------------------ Cancellations Sheet ------------------
    if os.path.exists(CANCELLATIONS_FILE):
        cancel_ws = load_workbook(CANCELLATIONS_FILE).active
        sheet = wb.create_sheet("Cancellations")
        for row in cancel_ws.iter_rows(values_only=True):
            sheet.append(row)

    # Save final summary Excel file
    wb.save(summary_path)

    # Send as Telegram document
    with open(summary_path, "rb") as f:
        bot.send_document(message.chat.id, InputFile(f, filename="ShiftSummary.xlsx"))

    # Cleanup temp file
    os.remove(summary_path)

def shift_reminder_loop():
    while True:
        now = datetime.now(pytz.timezone("Asia/Singapore"))
        today = now.date()

        # Load today's bookings
        if not os.path.exists(BOOKINGS_FILE):
            time.sleep(60)
            continue

        wb = load_workbook(BOOKINGS_FILE)
        ws = wb.active

        for shift_name, (start_str, _) in SHIFT_OPTIONS.items():
            shift_start_time = datetime.strptime(start_str, "%H:%M").time()
            shift_datetime = datetime.combine(today, shift_start_time)
            shift_datetime = pytz.timezone("Asia/Singapore").localize(shift_datetime)

            # Notify exactly 1 hour before the shift
            time_diff = (shift_datetime - now).total_seconds()
            if 3540 <= time_diff <= 3660:  # ~1 hour ±1 minute window
                students_in_shift = [
                    (row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[2] == today.strftime("%Y-%m-%d") and row[3] == shift_name
                ]

                if students_in_shift:
                    msg_lines = [f"*Shift Reminder: {shift_name} ({start_str})*", f"*Date:* {today.strftime('%Y-%m-%d')}"]
                    for sid, name in students_in_shift:
                        msg_lines.append(f"{name} (ID: {sid})")
                    message = "\n".join(msg_lines)

                    # Notify all logged-in users in the shift
                    for user_id, info in logged_in_users.items():
                        if (info["student_id"], info["name"]) in students_in_shift:
                            bot.send_message(user_id, message, parse_mode='Markdown')

        # Sleep 60 seconds before next check
        time.sleep(60)

# Run the bot
if __name__ == "__main__":
    print("Bot is running...")
    bot.infinity_polling()
